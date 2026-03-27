import json
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


def odaberi_fajl():
    root = tk.Tk()
    root.withdraw()
    putanja = filedialog.askopenfilename(
        title="Odaberi JSON fajl sa podacima",
        filetypes=[("JSON fajlovi", "*.json"), ("Svi fajlovi", "*.*")]
    )
    root.destroy()
    return putanja


def parsiraj_podatke(putanja_json):
    with open(putanja_json, encoding='utf-8') as f:
        data = json.load(f)

    ph = data['results'][0]['result']['data']['dsr']['DS'][0]['PH'][0]['DM0']

    rows = []
    for entry in ph:
        ts = entry['G0']
        datum = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).strftime('%d.%m.%Y')
        x = entry.get('X', [])
        kratko = None
        dugo = None

        if len(x) == 2:
            item0, item1 = x[0], x[1]
            if 'I' not in item0 and 'R' not in item0:
                kratko = float(item0['M0']) if item0.get('M0') is not None else None
            if 'I' in item1 and 'R' not in item1:
                dugo = float(item1['M0']) if item1.get('M0') is not None else None
            elif 'I' not in item1 and 'R' not in item1:
                dugo = float(item1['M0']) if item1.get('M0') is not None else None
        elif len(x) == 1:
            item = x[0]
            if 'R' in item:
                pass
            elif 'I' in item and item['I'] == 1:
                dugo = float(item['M0']) if item.get('M0') is not None else None
            else:
                kratko = float(item['M0']) if item.get('M0') is not None else None

        rows.append((datum, kratko, dugo))

    return rows


def napravi_excel(rows, putanja_xlsx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prinosi na kapital"

    # Stilovi
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E79')
    data_font = Font(name='Arial', size=10)
    pct_format = '0.00%'
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Zaglavlje
    for col, naziv in [('A', 'Datum'), ('B', 'Kratkoročne'), ('C', 'Dugoročne')]:
        cell = ws[f'{col}1']
        cell.value = naziv
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Podaci
    for i, (datum, kratko, dugo) in enumerate(rows, start=2):
        row_fill = PatternFill('solid', start_color='F2F2F2' if i % 2 == 0 else 'FFFFFF')

        ws[f'A{i}'] = datum
        ws[f'A{i}'].font = data_font
        ws[f'A{i}'].alignment = center
        ws[f'A{i}'].border = border
        ws[f'A{i}'].fill = row_fill

        ws[f'B{i}'] = kratko if kratko is not None else ''
        ws[f'B{i}'].font = data_font
        ws[f'B{i}'].alignment = center
        ws[f'B{i}'].border = border
        ws[f'B{i}'].fill = row_fill
        if kratko is not None:
            ws[f'B{i}'].number_format = pct_format

        ws[f'C{i}'] = dugo if dugo is not None else ''
        ws[f'C{i}'].font = data_font
        ws[f'C{i}'].alignment = center
        ws[f'C{i}'].border = border
        ws[f'C{i}'].fill = row_fill
        if dugo is not None:
            ws[f'C{i}'].number_format = pct_format

    # Širine kolona i zamrzavanje zaglavlja
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.freeze_panes = 'A2'

    wb.save(putanja_xlsx)


def main():
    print("=== Konvertor JSON → Excel (Prinosi na kapital) ===\n")

    putanja_json = odaberi_fajl()
    if not putanja_json:
        print("Nije odabran nijedan fajl. Program se završava.")
        return

    print(f"Odabran fajl: {putanja_json}")
    print("Parsiranje podataka...")

    try:
        rows = parsiraj_podatke(putanja_json)
    except Exception as e:
        messagebox.showerror("Greška", f"Greška pri čitanju JSON fajla:\n{e}")
        print(f"Greška: {e}")
        return

    # Sačuvaj xlsx u isti folder kao JSON, sa istim imenom
    folder = os.path.dirname(putanja_json)
    ime_bez_ext = os.path.splitext(os.path.basename(putanja_json))[0]
    putanja_xlsx = os.path.join(folder, f"{ime_bez_ext}_prinosi.xlsx")

    print(f"Kreiranje Excel fajla ({len(rows)} redova)...")

    try:
        napravi_excel(rows, putanja_xlsx)
    except Exception as e:
        messagebox.showerror("Greška", f"Greška pri kreiranju Excel fajla:\n{e}")
        print(f"Greška: {e}")
        return

    poruka = f"Uspješno sačuvano!\n\nFajl: {putanja_xlsx}\nBroj redova: {len(rows)}"
    print(f"\n{poruka}")

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Gotovo!", poruka)
    root.destroy()


if __name__ == '__main__':
    main()
